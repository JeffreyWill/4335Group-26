"""
Запусти этот скрипт чтобы создать тестовый файл 5.xlsx
Команда: python create_5xlsx.py
"""
import subprocess, sys
subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
ws = wb.active
ws.title = "Сотрудники"

headers = ["Логин", "Пароль", "Роль", "ФИО", "E-mail"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(fill_type="solid", fgColor="2C3E50")
    cell.alignment = Alignment(horizontal="center")

data = [
    ("ivanov",   "pass123",   "Администратор", "Иванов Иван Иванович",     "ivanov@corp.ru"),
    ("petrov",   "qwerty",    "Менеджер",      "Петров Пётр Петрович",     "petrov@corp.ru"),
    ("sidorov",  "abc456",    "Менеджер",      "Сидоров Сидор Сидорович",  "sidorov@corp.ru"),
    ("kozlov",   "xyz789",    "Оператор",      "Козлов Козьма Козьмич",    "kozlov@corp.ru"),
    ("morozov",  "secret1",   "Оператор",      "Морозов Мороз Морозович",  "morozov@corp.ru"),
    ("novikov",  "admin2024", "Администратор", "Новиков Новик Новикович",  "novikov@corp.ru"),
    ("sokolov",  "oper1",     "Оператор",      "Соколов Сокол Соколович",  "sokolov@corp.ru"),
    ("volkov",   "mgr2024",   "Менеджер",      "Волков Вол Волкович",      "volkov@corp.ru"),
    ("lebedev",  "leb555",    "Администратор", "Лебедев Лебедь Лебедович", "lebedev@corp.ru"),
    ("popov",    "pop999",    "Оператор",      "Попов Поп Попович",        "popov@corp.ru"),
]

for ri, row in enumerate(data, 2):
    for ci, val in enumerate(row, 1):
        ws.cell(row=ri, column=ci, value=val)

for col in ws.columns:
    ws.column_dimensions[col[0].column_letter].width = \
        max(len(str(c.value or "")) for c in col) + 4

wb.save("5.xlsx")
print("✅ Файл 5.xlsx успешно создан!")
